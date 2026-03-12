from __future__ import annotations

import shutil
from collections.abc import Iterator
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl.worksheet.table import Table, TableStyleInfo

from orchestra_agent.mcp_server.excel_service import ExcelToolError, ExcelWorkspaceService

openpyxl = pytest.importorskip("openpyxl")
PILImage = pytest.importorskip("PIL.Image")


@pytest.fixture()
def workspace_dir() -> Iterator[Path]:
    base = Path(".tmp-tests") / uuid4().hex
    base.mkdir(parents=True, exist_ok=False)
    try:
        yield base
    finally:
        shutil.rmtree(base, ignore_errors=True)


def test_excel_service_reads_sums_writes_and_exports(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "sales.xlsx"
    output_path = workspace_dir / "summary.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["C1"] = "Amount"
    sheet["C2"] = 10
    sheet["C3"] = "20"
    sheet["C4"] = 30
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)

    opened = service.open_file("sales.xlsx")
    assert opened["sheet_names"] == ["Sheet1"]

    rows = service.read_sheet("sales.xlsx", "Sheet1")
    assert rows["row_count"] == 4
    assert rows["rows"][1]["C"] == 10

    total = service.calculate_sum("sales.xlsx", "Sheet1", "C")
    assert total["total"] == 60
    assert total["counted_cells"] == 3

    created = service.create_sheet("sales.xlsx", "Summary")
    assert created["created"] is True

    written = service.write_cells(
        "sales.xlsx",
        "Summary",
        {"A1": "Column", "B1": "Total", "A2": "C", "B2": 60},
    )
    assert written["written_cells"] == 4

    exported = service.save_file("sales.xlsx", "summary.xlsx")
    assert exported["output"] == "summary.xlsx"
    assert output_path.is_file()

    exported_workbook = openpyxl.load_workbook(output_path)
    try:
        assert exported_workbook["Summary"]["B2"].value == 60
    finally:
        exported_workbook.close()


def test_excel_service_creates_new_workbook_file(workspace_dir: Path) -> None:
    service = ExcelWorkspaceService(workspace_dir)

    created = service.create_file("output/new_report.xlsx", sheet="Data")

    assert created["file"] == "output/new_report.xlsx"
    assert created["sheet_names"] == ["Data"]
    assert created["created"] is True
    assert created["overwritten"] is False
    assert (workspace_dir / "output" / "new_report.xlsx").is_file()

    opened = service.open_file("output/new_report.xlsx")
    assert opened["sheet_names"] == ["Data"]


def test_excel_service_reads_specific_cells_and_greps_content(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "instructions.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Keyword"
    sheet["B2"] = "Review this program"
    sheet["C3"] = "会いう"
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)

    cells = service.read_cells("instructions.xlsx", "Sheet1", ["A1", "B2", "C3"])
    assert cells["cells"] == {
        "A1": "Keyword",
        "B2": "Review this program",
        "C3": "会いう",
    }

    grep = service.grep_cells("instructions.xlsx", "会いう")
    assert grep["matches"] == [{"sheet": "Sheet1", "cell": "C3", "value": "会いう"}]
    assert grep["truncated"] is False


def test_excel_service_lists_and_extracts_embedded_images(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "images.xlsx"
    image_path = workspace_dir / "sample.png"
    output_path = workspace_dir / "artifacts" / "extracted.png"

    PILImage.new("RGB", (12, 12), color="red").save(image_path)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    embedded = openpyxl.drawing.image.Image(str(image_path))
    sheet.add_image(embedded, "B2")
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)

    listed = service.list_images("images.xlsx")
    assert listed["images"] == [
        {
            "sheet": "Sheet1",
            "image_index": 1,
            "anchor_cell": "B2",
            "extension": ".png",
            "zip_path": "xl/media/image1.png",
        }
    ]

    extracted = service.extract_image(
        "images.xlsx",
        sheet="Sheet1",
        image_index=1,
        output="artifacts/extracted.png",
    )
    assert extracted["output"] == "artifacts/extracted.png"
    assert output_path.is_file()
    assert output_path.read_bytes() == image_path.read_bytes()


def test_excel_service_blocks_paths_outside_workspace(workspace_dir: Path) -> None:
    service = ExcelWorkspaceService(workspace_dir)

    with pytest.raises(PermissionError):
        service.open_file("../outside.xlsx")


def test_excel_service_inspects_ranges_tables_and_sources(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "report.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Name"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Alice"
    sheet["B2"] = 10
    sheet["A3"] = "Bob"
    sheet["B3"] = 20
    table = Table(displayName="SalesTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)

    listed = service.list_sources()
    assert listed["sources"][0]["source_id"] == "local_workspace"

    resolved = service.resolve_workbook("local_workspace", path="report.xlsx")
    workbook_ref = resolved["workbook_ref"]

    inspected = service.inspect_workbook(
        workbook_ref,
        include_sheet_stats=True,
        include_tables=True,
    )
    assert inspected["filename"] == "report.xlsx"
    assert inspected["sheets"][0]["used_range"] == "A1:B3"
    assert inspected["tables"] == [
        {
            "sheet": "Data",
            "table_name": "SalesTable",
            "address": "A1:B3",
            "column_count": 2,
            "row_count": 2,
        }
    ]

    read_range = service.read_range(workbook_ref, sheet="Data", range="A1:B3")
    assert read_range["values"] == [["Name", "Amount"], ["Alice", 10], ["Bob", 20]]

    read_table = service.read_table(workbook_ref, table_name="SalesTable")
    assert read_table["headers"] == ["Name", "Amount"]
    assert read_table["rows"] == [["Alice", 10], ["Bob", 20]]


def test_excel_service_edit_session_commit_creates_backup_and_updates_workbook(
    workspace_dir: Path,
) -> None:
    workbook_path = workspace_dir / "sales.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Name"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Alice"
    sheet["B2"] = 10
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)
    workbook_ref = service.resolve_workbook("local_workspace", path="sales.xlsx")["workbook_ref"]

    opened = service.open_edit_session(workbook_ref)
    session_id = opened["session_id"]

    created = service.stage_create_sheet(session_id, "Summary")
    assert "operation_id" in created

    updated = service.stage_update_cells(
        session_id,
        sheet="Summary",
        start_cell="A1",
        values=[["Metric", "Value"], ["Rows", 1]],
    )
    assert updated["affected_range"] == "A1:B2"

    appended = service.stage_append_rows(
        session_id,
        sheet="Sheet1",
        rows=[["Bob", 20]],
    )
    assert appended["affected_table_or_range"] == "A3:B3"

    preview = service.preview_edit_session(session_id, detail_level="cell_level")["preview"]
    assert sorted(preview["changed_sheets"]) == ["Sheet1", "Summary"]
    assert "CREATE_SHEET" in preview["potential_risk_flags"]

    validation = service.validate_edit_session(session_id)
    assert validation["valid"] is True

    committed = service.commit_edit_session(session_id, commit_message="add summary")
    assert committed["backup_ref"] is not None

    committed_workbook = openpyxl.load_workbook(workbook_path)
    try:
        assert committed_workbook["Summary"]["B2"].value == 1
        assert committed_workbook["Sheet1"]["A3"].value == "Bob"
        assert committed_workbook["Sheet1"]["B3"].value == 20
    finally:
        committed_workbook.close()

    backups = service.list_backups("local_workspace", target="sales.xlsx")
    assert len(backups["backups"]) == 1
    assert backups["backups"][0]["original_path"] == "sales.xlsx"


def test_excel_service_detects_conflict_before_commit(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "sales.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Amount"
    sheet["A2"] = 10
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)
    workbook_ref = service.resolve_workbook("local_workspace", path="sales.xlsx")["workbook_ref"]
    session_id = service.open_edit_session(workbook_ref)["session_id"]

    service.stage_update_cells(
        session_id,
        sheet="Sheet1",
        start_cell="A2",
        values=[[20]],
    )
    _ = service.preview_edit_session(session_id)

    conflicting = openpyxl.load_workbook(workbook_path)
    try:
        conflicting["Sheet1"]["A2"] = 99
        conflicting.save(workbook_path)
    finally:
        conflicting.close()

    validation = service.validate_edit_session(session_id)
    assert validation["valid"] is False
    assert validation["errors"][0]["code"] == "CONFLICT_DETECTED"

    with pytest.raises(ExcelToolError) as exc_info:
        service.commit_edit_session(session_id)
    assert exc_info.value.code == "CONFLICT_DETECTED"


def test_excel_service_restores_backup(workspace_dir: Path) -> None:
    workbook_path = workspace_dir / "restore.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Status"
    sheet["A2"] = "before"
    workbook.save(workbook_path)
    workbook.close()

    service = ExcelWorkspaceService(workspace_dir)
    workbook_ref = service.resolve_workbook("local_workspace", path="restore.xlsx")["workbook_ref"]
    session_id = service.open_edit_session(workbook_ref)["session_id"]

    service.stage_update_cells(
        session_id,
        sheet="Sheet1",
        start_cell="A2",
        values=[["after"]],
    )
    _ = service.preview_edit_session(session_id)
    _ = service.validate_edit_session(session_id)
    committed = service.commit_edit_session(session_id)
    backup_ref = committed["backup_ref"]
    assert backup_ref is not None

    modified = openpyxl.load_workbook(workbook_path)
    try:
        modified["Sheet1"]["A2"] = "mutated"
        modified.save(workbook_path)
    finally:
        modified.close()

    restored = service.restore_backup(backup_ref)
    assert restored["restore_result"]["restored"] is True

    restored_workbook = openpyxl.load_workbook(workbook_path)
    try:
        assert restored_workbook["Sheet1"]["A2"].value == "before"
    finally:
        restored_workbook.close()
