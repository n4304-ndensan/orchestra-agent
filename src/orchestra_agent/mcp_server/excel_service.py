from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import time as time_module
import xml.etree.ElementTree as ET
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, timedelta
from datetime import time as time_value
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Any, Literal
from uuid import uuid4
from zipfile import ZipFile

from orchestra_agent.mcp_server.excel_config import (
    ExcelServerConfig,
    ExcelSourceProfile,
    load_excel_server_config,
)
from orchestra_agent.mcp_server.logging_utils import get_mcp_logger, log_event

logger = get_mcp_logger(__name__)

type SessionState = Literal[
    "CREATED",
    "STAGING",
    "PREVIEWED",
    "VALIDATED",
    "COMMITTED",
    "CANCELED",
    "FAILED",
]
type PreviewDetailLevel = Literal["summary", "detailed", "cell_level"]
type ValueRenderMode = Literal["raw", "formatted", "formula"]


class ExcelToolError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        detail: Mapping[str, Any] | None = None,
        retriable: bool = False,
        suggested_action: str | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.detail = dict(detail or {})
        self.retriable = retriable
        self.suggested_action = suggested_action

    def to_dict(self) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "code": self.code,
            "message": self.message,
            "detail": dict(self.detail),
            "retriable": self.retriable,
        }
        if self.suggested_action is not None:
            payload["suggested_action"] = self.suggested_action
        return payload

    def __str__(self) -> str:
        return self.message


@dataclass(slots=True)
class CellChange:
    sheet: str
    cell: str
    old_value: Any
    new_value: Any

    @property
    def formula(self) -> bool:
        return _is_formula_value(self.new_value)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "cell": self.cell,
            "old_value": _serialize_value(self.old_value),
            "new_value": _serialize_value(self.new_value),
            "formula": self.formula,
        }


@dataclass(slots=True)
class StagedOperation:
    operation_id: str
    operation_type: str
    sheet: str | None = None
    affected_range: str | None = None
    warnings: list[str] = field(default_factory=list)
    risk_flags: list[str] = field(default_factory=list)
    cell_changes: list[CellChange] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_preview_dict(self, detail_level: PreviewDetailLevel) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "operation_id": self.operation_id,
            "operation_type": self.operation_type,
            "sheet": self.sheet,
            "affected_range": self.affected_range,
            "warnings": list(self.warnings),
            "risk_flags": list(self.risk_flags),
        }
        if self.metadata:
            payload["metadata"] = dict(self.metadata)
        if detail_level in ("detailed", "cell_level"):
            payload["cell_change_count"] = len(self.cell_changes)
        if detail_level == "cell_level":
            payload["cell_changes"] = [change.to_dict() for change in self.cell_changes]
        return payload


@dataclass(slots=True)
class EditSession:
    session_id: str
    source_id: str
    source_mode: str
    target_ref: dict[str, Any]
    opened_at: datetime
    expires_at: datetime
    actor: str
    read_only: bool
    base_version: str
    base_etag: str | None
    base_hash: str
    base_size: int
    base_modified_at: str
    temp_workbook_path: Path
    staged_operations: list[StagedOperation] = field(default_factory=list)
    preview_summary: dict[str, Any] | None = None
    validation_summary: dict[str, Any] | None = None
    backup_ref: dict[str, Any] | None = None
    commit_result: dict[str, Any] | None = None
    audit_ref: str | None = None
    state: SessionState = "CREATED"
    last_accessed_at: datetime = field(default_factory=lambda: datetime.now(UTC))

    def touch(self) -> None:
        self.last_accessed_at = datetime.now(UTC)


class ExcelWorkspaceService:
    _main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    _rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    _pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    _xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    _a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def __init__(
        self,
        workspace_root: Path,
        config: ExcelServerConfig | None = None,
    ) -> None:
        self._workspace_root = Path(workspace_root).resolve()
        self._config = config or load_excel_server_config(self._workspace_root)
        self._sources = self._normalize_sources(self._config.sources)
        self._sessions: dict[str, EditSession] = {}
        self._audit_file = self._resolve_inside_workspace(
            self._config.logging.audit_file
            if self._config.logging is not None
            else self._workspace_root / ".orchestra_state" / "audit" / "excel_workspace_mcp.jsonl"
        )
        self._audit_file.parent.mkdir(parents=True, exist_ok=True)
        log_event(
            logger,
            "excel_service_initialized",
            workspace_root=self._workspace_root,
            source_count=len(self._sources),
            audit_file=self._audit_file,
        )

    @property
    def workspace_root(self) -> Path:
        return self._workspace_root

    def list_sources(self, include_disabled: bool = False) -> dict[str, Any]:
        started = time_module.perf_counter()
        sources = [
            source.to_public_dict()
            for source in self._sources.values()
            if include_disabled or source.enabled
        ]
        result = {"sources": sources}
        self._audit(
            operation="list_sources",
            result="success",
            duration_ms=_duration_ms(started),
        )
        return result

    def find_workbooks(
        self,
        source_id: str,
        query: str = "",
        *,
        path_prefix: str | None = None,
        recursive: bool = True,
        limit: int | None = None,
        extension_filter: Sequence[str] | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source = self._require_local_source(source_id)
        effective_limit = limit if limit is not None else self._config.limits.search_result_limit
        if effective_limit <= 0:
            self._raise_error("VALIDATION_FAILED", "limit must be greater than zero.")
        search_root = source.workspace_root
        if path_prefix:
            search_root = self._resolve_path_inside_root(path_prefix, source.workspace_root)
            if not search_root.exists():
                result = {"workbook_refs": []}
                self._audit(
                    operation="find_workbooks",
                    result="success",
                    source_id=source.source_id,
                    target_file=path_prefix,
                    duration_ms=_duration_ms(started),
                )
                return result
        normalized_query = query.strip().lower()
        allowed_extensions = self._normalize_extension_filter(extension_filter, source)
        matches: list[dict[str, Any]] = []
        iterator = search_root.rglob("*") if recursive else search_root.glob("*")
        for candidate in iterator:
            if not candidate.is_file():
                continue
            if candidate.suffix.lower() not in allowed_extensions:
                continue
            relative_path = candidate.relative_to(source.workspace_root).as_posix()
            if normalized_query and normalized_query not in relative_path.lower():
                continue
            matches.append(self._build_workbook_ref(source, candidate, include_hash=False))
            if len(matches) >= effective_limit:
                break
        result = {"workbook_refs": matches}
        self._audit(
            operation="find_workbooks",
            result="success",
            source_id=source.source_id,
            target_file=path_prefix,
            duration_ms=_duration_ms(started),
        )
        return result

    def resolve_workbook(
        self,
        source_id: str,
        *,
        path: str | None = None,
        remote_ref: Mapping[str, Any] | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        if path is None and remote_ref is None:
            self._raise_error(
                "WORKBOOK_NOT_FOUND",
                "resolve_workbook requires a path or remote_ref.",
            )
        source = self._require_source(source_id)
        if source.source_type != "local_workspace":
            self._raise_remote_not_supported(source)
        if path is None:
            candidate_path = remote_ref.get("path") if remote_ref is not None else None
            if not isinstance(candidate_path, str) or not candidate_path.strip():
                self._raise_error(
                    "WORKBOOK_NOT_FOUND",
                    "Local resolve_workbook requires a string path.",
                )
            path = candidate_path
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook_ref = self._build_workbook_ref(source, workbook_path)
        result = {"workbook_ref": workbook_ref}
        self._audit(
            operation="resolve_workbook",
            result="success",
            source_id=source.source_id,
            target_file=workbook_ref["path"],
            duration_ms=_duration_ms(started),
        )
        return result

    def inspect_workbook(
        self,
        workbook_ref: Mapping[str, Any] | str,
        *,
        include_sheet_stats: bool = True,
        include_tables: bool = False,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source, workbook_path, ref = self._resolve_workbook_input(workbook_ref)
        workbook = self._load_workbook(workbook_path, data_only=False)
        try:
            sheets: list[dict[str, Any]] = []
            tables: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                sheet_entry: dict[str, Any] = {
                    "name": worksheet.title,
                    "hidden": worksheet.sheet_state != "visible",
                }
                if include_sheet_stats:
                    sheet_entry["used_range"] = worksheet.calculate_dimension()
                    sheet_entry["row_count"] = worksheet.max_row
                    sheet_entry["column_count"] = worksheet.max_column
                    sheet_entry["table_count"] = len(worksheet.tables)
                sheets.append(sheet_entry)
                if include_tables:
                    for table in worksheet.tables.values():
                        min_col, min_row, max_col, max_row = self._range_boundaries(table.ref)
                        tables.append(
                            {
                                "sheet": worksheet.title,
                                "table_name": table.displayName,
                                "address": table.ref,
                                "column_count": max_col - min_col + 1,
                                "row_count": max_row - min_row,
                            }
                        )

            named_ranges = [
                {"name": defined_name.name, "value": defined_name.attr_text}
                for defined_name in workbook.defined_names.values()
            ]
            result = {
                "workbook_ref": ref,
                "filename": workbook_path.name,
                "size": workbook_path.stat().st_size,
                "modified_at": _file_mtime_iso(workbook_path),
                "hash": _sha256_file(workbook_path),
                "extension": workbook_path.suffix.lower(),
                "sheets": sheets,
                "tables": tables,
                "named_ranges": named_ranges,
            }
        finally:
            workbook.close()
        self._audit(
            operation="inspect_workbook",
            result="success",
            source_id=source.source_id,
            target_file=ref["path"],
            duration_ms=_duration_ms(started),
        )
        return result

    def list_sheets(self, workbook_ref: Mapping[str, Any] | str) -> dict[str, Any]:
        inspected = self.inspect_workbook(
            workbook_ref,
            include_sheet_stats=False,
            include_tables=False,
        )
        return {"sheets": inspected["sheets"]}

    def read_range(
        self,
        workbook_ref: Mapping[str, Any] | str,
        sheet: str,
        range: str,
        *,
        value_render_mode: ValueRenderMode = "raw",
        max_cells: int | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source, workbook_path, ref = self._resolve_workbook_input(workbook_ref)
        effective_max = (
            max_cells
            if max_cells is not None
            else self._config.limits.read_range_max_cells
        )
        min_col, min_row, max_col, max_row = self._range_boundaries(range)
        cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
        if cell_count > effective_max:
            self._raise_error(
                "RANGE_INVALID",
                f"Requested range exceeds the max cell limit of {effective_max}.",
                detail={"range": range, "max_cells": effective_max, "requested_cells": cell_count},
            )
        workbook = self._load_workbook(
            workbook_path,
            data_only=value_render_mode != "formula",
        )
        try:
            worksheet = self._get_sheet(workbook, sheet)
            values: list[list[Any]] = []
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                rendered_row = [
                    self._render_cell_value(cell.value, value_render_mode) for cell in row
                ]
                values.append(rendered_row)
        finally:
            workbook.close()
        result = {
            "values": values,
            "address": range.upper(),
            "row_count": max_row - min_row + 1,
            "col_count": max_col - min_col + 1,
        }
        self._audit(
            operation="read_range",
            result="success",
            source_id=source.source_id,
            source_mode=source.default_mode,
            target_file=ref["path"],
            target_sheet=sheet,
            target_range=range,
            duration_ms=_duration_ms(started),
        )
        return result

    def read_table(
        self,
        workbook_ref: Mapping[str, Any] | str,
        table_name: str,
        *,
        sheet: str | None = None,
        max_rows: int | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source, workbook_path, ref = self._resolve_workbook_input(workbook_ref)
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            worksheet, table = self._find_table(workbook, table_name, sheet=sheet)
            min_col, min_row, max_col, max_row = self._range_boundaries(table.ref)
            data_row_count = max(0, max_row - min_row)
            effective_max = (
                max_rows
                if max_rows is not None
                else self._config.limits.read_table_max_rows
            )
            if data_row_count > effective_max:
                self._raise_error(
                    "VALIDATION_FAILED",
                    f"Table '{table_name}' exceeds the max row limit of {effective_max}.",
                    detail={"row_count": data_row_count, "max_rows": effective_max},
                )
            headers = [
                _serialize_value(worksheet.cell(row=min_row, column=column).value)
                for column in range(min_col, max_col + 1)
            ]
            rows: list[list[Any]] = []
            for row_index in range(min_row + 1, max_row + 1):
                rows.append(
                    [
                        _serialize_value(worksheet.cell(row=row_index, column=column).value)
                        for column in range(min_col, max_col + 1)
                    ]
                )
        finally:
            workbook.close()
        result = {
            "headers": headers,
            "rows": rows,
            "address": table.ref,
            "row_count": len(rows),
        }
        self._audit(
            operation="read_table",
            result="success",
            source_id=source.source_id,
            source_mode=source.default_mode,
            target_file=ref["path"],
            target_sheet=worksheet.title,
            target_range=table.ref,
            duration_ms=_duration_ms(started),
        )
        return result

    def search_workbook_text(
        self,
        workbook_ref: Mapping[str, Any] | str,
        pattern: str,
        *,
        match_case: bool = False,
        exact: bool = False,
        max_results: int | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source, workbook_path, ref = self._resolve_workbook_input(workbook_ref)
        effective_max = (
            max_results
            if max_results is not None
            else self._config.limits.search_result_limit
        )
        if effective_max <= 0:
            self._raise_error("VALIDATION_FAILED", "max_results must be greater than zero.")
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            matches: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        if not _matches_text(
                            text=str(cell.value),
                            pattern=pattern,
                            case_sensitive=match_case,
                            regex=False,
                            exact=exact,
                        ):
                            continue
                        matches.append(
                            {
                                "sheet": worksheet.title,
                                "cell": cell.coordinate,
                                "value": _serialize_value(cell.value),
                            }
                        )
                        if len(matches) >= effective_max:
                            result = {"matches": matches, "truncated": True}
                            self._audit(
                                operation="search_workbook_text",
                                result="success",
                                source_id=source.source_id,
                                source_mode=source.default_mode,
                                target_file=ref["path"],
                                duration_ms=_duration_ms(started),
                            )
                            return result
        finally:
            workbook.close()
        result = {"matches": matches, "truncated": False}
        self._audit(
            operation="search_workbook_text",
            result="success",
            source_id=source.source_id,
            source_mode=source.default_mode,
            target_file=ref["path"],
            duration_ms=_duration_ms(started),
        )
        return result

    def open_edit_session(
        self,
        workbook_ref: Mapping[str, Any] | str,
        *,
        source_mode: str | None = None,
        read_only: bool = False,
        backup_policy: Mapping[str, Any] | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        source, workbook_path, ref = self._resolve_workbook_input(workbook_ref)
        if source.source_type != "local_workspace":
            self._raise_remote_not_supported(source)
        if source.read_only and not read_only:
            self._raise_error(
                "SOURCE_READ_ONLY",
                f"Source '{source.source_id}' is read-only.",
            )
        fingerprint = self._file_fingerprint(workbook_path)
        session_id = uuid4().hex
        temp_root = self._resolve_path_inside_root(session_id, source.temp_root)
        temp_root.mkdir(parents=True, exist_ok=False)
        temp_workbook_path = temp_root / workbook_path.name
        shutil.copy2(workbook_path, temp_workbook_path)
        now = datetime.now(UTC)
        session = EditSession(
            session_id=session_id,
            source_id=source.source_id,
            source_mode=source_mode or source.default_mode,
            target_ref=ref,
            opened_at=now,
            expires_at=now + timedelta(seconds=self._config.limits.hard_timeout_sec),
            actor=self._actor(),
            read_only=read_only,
            base_version=fingerprint["hash"],
            base_etag=None,
            base_hash=fingerprint["hash"],
            base_size=int(fingerprint["size"]),
            base_modified_at=str(fingerprint["modified_at"]),
            temp_workbook_path=temp_workbook_path,
            state="CREATED",
        )
        if backup_policy is not None:
            session.backup_ref = {"requested_policy": dict(backup_policy)}
        session.audit_ref = self._audit_file.relative_to(self._workspace_root).as_posix()
        self._sessions[session_id] = session
        result = {
            "session_id": session_id,
            "base_version": session.base_version,
            "expires_at": session.expires_at.isoformat(),
        }
        self._audit(
            operation="open_edit_session",
            result="success",
            source_id=source.source_id,
            source_mode=session.source_mode,
            target_file=ref["path"],
            duration_ms=_duration_ms(started),
        )
        return result

    def stage_update_cells(  # noqa: C901
        self,
        session_id: str,
        sheet: str,
        start_cell: str,
        values: Sequence[Sequence[Any]],
        *,
        write_mode: str = "overwrite",
        expected_existing_values: Sequence[Sequence[Any]] | None = None,
        allow_formula: bool | None = None,
    ) -> dict[str, Any]:
        session = self._require_session(session_id, writable=True)
        matrix = _normalize_matrix(values, field_name="values")
        if expected_existing_values is not None:
            expected_matrix = _normalize_matrix(
                expected_existing_values,
                field_name="expected_existing_values",
            )
            if _matrix_shape(expected_matrix) != _matrix_shape(matrix):
                self._raise_error(
                    "VALIDATION_FAILED",
                    "expected_existing_values must have the same dimensions as values.",
                )
        else:
            expected_matrix = None

        total_cells = len(matrix) * len(matrix[0])
        if total_cells > self._config.limits.max_cells_per_update:
            self._raise_error(
                "VALIDATION_FAILED",
                f"update exceeds the max cell limit of {self._config.limits.max_cells_per_update}.",
            )
        if write_mode not in {"overwrite", "only_blank", "reject_if_nonblank"}:
            self._raise_error("VALIDATION_FAILED", f"Unsupported write_mode '{write_mode}'.")

        workbook = self._load_workbook(session.temp_workbook_path, data_only=False)
        try:
            worksheet = self._get_sheet(workbook, sheet)
            start_column_index, start_row_index = self._cell_indices(start_cell)
            warnings: list[str] = []
            risk_flags: list[str] = []
            changes: list[CellChange] = []
            for row_offset, row_values in enumerate(matrix):
                for column_offset, new_value in enumerate(row_values):
                    row_index = start_row_index + row_offset
                    column_index = start_column_index + column_offset
                    cell = worksheet.cell(row=row_index, column=column_index)
                    self._ensure_not_merged_overlap(worksheet, cell.coordinate)
                    old_value = cell.value
                    if expected_matrix is not None:
                        expected_value = expected_matrix[row_offset][column_offset]
                        if _serialize_value(old_value) != _serialize_value(expected_value):
                            self._raise_error(
                                "VALIDATION_FAILED",
                                f"Existing value mismatch at {cell.coordinate}.",
                                detail={
                                    "sheet": sheet,
                                    "cell": cell.coordinate,
                                    "expected": _serialize_value(expected_value),
                                    "actual": _serialize_value(old_value),
                                },
                            )
                    if write_mode == "only_blank" and not _is_blank(old_value):
                        warnings.append(f"Skipped non-blank cell {cell.coordinate}.")
                        continue
                    if write_mode == "reject_if_nonblank" and not _is_blank(old_value):
                        self._raise_error(
                            "VALIDATION_FAILED",
                            f"Cell {cell.coordinate} is not blank.",
                            detail={"sheet": sheet, "cell": cell.coordinate},
                        )
                    if _is_formula_value(new_value):
                        allow_formula_write = (
                            self._config.policies.allow_formula_write
                            if allow_formula is None
                            else allow_formula
                        )
                        if not allow_formula_write:
                            self._raise_error(
                                "VALIDATION_FAILED",
                                f"Formula writes are not allowed for cell {cell.coordinate}.",
                            )
                        risk_flags.append("FORMULA_WRITE")
                    if (
                        not _is_blank(old_value)
                        and _serialize_value(old_value) != _serialize_value(new_value)
                    ):
                        risk_flags.append("OVERWRITE_NONBLANK")
                    if _serialize_value(old_value) == _serialize_value(new_value):
                        continue
                    cell.value = new_value
                    changes.append(
                        CellChange(
                            sheet=sheet,
                            cell=cell.coordinate,
                            old_value=old_value,
                            new_value=new_value,
                        )
                    )
            workbook.save(session.temp_workbook_path)
        finally:
            workbook.close()

        operation_id = uuid4().hex
        affected_range = _matrix_range(start_cell, row_count=len(matrix), col_count=len(matrix[0]))
        operation = StagedOperation(
            operation_id=operation_id,
            operation_type="update_cells",
            sheet=sheet,
            affected_range=affected_range,
            warnings=_unique_list(warnings),
            risk_flags=_unique_list(risk_flags),
            cell_changes=changes,
            metadata={"write_mode": write_mode},
        )
        self._record_operation(session, operation)
        return {
            "operation_id": operation_id,
            "affected_range": affected_range,
            "warnings": list(operation.warnings),
        }

    def stage_append_rows(
        self,
        session_id: str,
        sheet: str,
        rows: Sequence[Sequence[Any]],
        *,
        table_name: str | None = None,
        anchor_range: str | None = None,
        schema_policy: str = "strict",
    ) -> dict[str, Any]:
        session = self._require_session(session_id, writable=True)
        matrix = _normalize_matrix(rows, field_name="rows")
        if len(matrix) > self._config.limits.max_rows_per_append:
            self._raise_error(
                "VALIDATION_FAILED",
                f"append exceeds the max row limit of {self._config.limits.max_rows_per_append}.",
            )
        workbook = self._load_workbook(session.temp_workbook_path, data_only=False)
        try:
            worksheet = self._get_sheet(workbook, sheet)
            warnings: list[str] = []
            risk_flags: list[str] = []
            changes: list[CellChange] = []
            if table_name is not None:
                table = worksheet.tables.get(table_name)
                if table is None:
                    self._raise_error(
                        "TABLE_NOT_FOUND",
                        f"Table '{table_name}' does not exist on sheet '{sheet}'.",
                    )
                min_col, min_row, max_col, max_row = self._range_boundaries(table.ref)
                expected_width = max_col - min_col + 1
                if schema_policy == "strict" and any(len(row) != expected_width for row in matrix):
                    self._raise_error(
                        "VALIDATION_FAILED",
                        f"Rows do not match the schema width of table '{table_name}'.",
                        detail={"expected_width": expected_width},
                    )
                start_row = max_row + 1
                start_col = min_col
                risk_flags.append("TABLE_RESIZE")
            else:
                if anchor_range is not None:
                    min_col, _, _, max_row = self._range_boundaries(anchor_range)
                    start_row = max_row + 1
                    start_col = min_col
                else:
                    start_row = 1 if _worksheet_is_empty(worksheet) else worksheet.max_row + 1
                    start_col = 1
                expected_width = len(matrix[0])
                if schema_policy == "strict" and any(len(row) != expected_width for row in matrix):
                    self._raise_error(
                        "VALIDATION_FAILED",
                        "All appended rows must have the same width.",
                    )

            for row_offset, row_values in enumerate(matrix):
                for column_offset, new_value in enumerate(row_values):
                    row_index = start_row + row_offset
                    column_index = start_col + column_offset
                    cell = worksheet.cell(row=row_index, column=column_index)
                    self._ensure_not_merged_overlap(worksheet, cell.coordinate)
                    old_value = cell.value
                    cell.value = new_value
                    changes.append(
                        CellChange(
                            sheet=sheet,
                            cell=cell.coordinate,
                            old_value=old_value,
                            new_value=new_value,
                        )
                    )

            end_col = start_col + len(matrix[0]) - 1
            end_row = start_row + len(matrix) - 1
            affected_range = _cell_range(
                start_col=start_col,
                start_row=start_row,
                end_col=end_col,
                end_row=end_row,
            )
            if table_name is not None:
                table.ref = _cell_range(
                    start_col=min_col,
                    start_row=min_row,
                    end_col=max_col,
                    end_row=end_row,
                )

            workbook.save(session.temp_workbook_path)
        finally:
            workbook.close()

        operation_id = uuid4().hex
        operation = StagedOperation(
            operation_id=operation_id,
            operation_type="append_rows",
            sheet=sheet,
            affected_range=affected_range,
            warnings=warnings,
            risk_flags=_unique_list(risk_flags),
            cell_changes=changes,
            metadata={
                "table_name": table_name,
                "anchor_range": anchor_range,
                "row_count": len(matrix),
            },
        )
        self._record_operation(session, operation)
        return {
            "operation_id": operation_id,
            "affected_table_or_range": table_name or affected_range,
            "warnings": list(operation.warnings),
        }

    def stage_create_sheet(
        self,
        session_id: str,
        new_sheet_name: str,
        *,
        template_sheet: str | None = None,
    ) -> dict[str, Any]:
        session = self._require_session(session_id, writable=True)
        sheet_name = new_sheet_name.strip()
        if not sheet_name:
            self._raise_error("VALIDATION_FAILED", "new_sheet_name must be non-empty.")
        workbook = self._load_workbook(session.temp_workbook_path, data_only=False)
        try:
            if sheet_name in workbook.sheetnames:
                self._raise_error(
                    "VALIDATION_FAILED",
                    f"Worksheet '{sheet_name}' already exists.",
                )
            if template_sheet is not None:
                source_sheet = self._get_sheet(workbook, template_sheet)
                copied = workbook.copy_worksheet(source_sheet)
                copied.title = sheet_name
            else:
                workbook.create_sheet(title=sheet_name)
            workbook.save(session.temp_workbook_path)
        finally:
            workbook.close()

        operation_id = uuid4().hex
        operation = StagedOperation(
            operation_id=operation_id,
            operation_type="create_sheet",
            sheet=sheet_name,
            affected_range=sheet_name,
            risk_flags=["CREATE_SHEET"],
            metadata={"template_sheet": template_sheet},
        )
        self._record_operation(session, operation)
        return {"operation_id": operation_id}

    def preview_edit_session(
        self,
        session_id: str,
        *,
        detail_level: PreviewDetailLevel = "summary",
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        session = self._require_session(session_id)
        if detail_level not in {"summary", "detailed", "cell_level"}:
            self._raise_error("VALIDATION_FAILED", f"Unsupported detail_level '{detail_level}'.")

        changed_sheets = _unique_list(
            [operation.sheet for operation in session.staged_operations if operation.sheet]
        )
        changed_ranges = _unique_list(
            [
                operation.affected_range
                for operation in session.staged_operations
                if operation.affected_range
            ]
        )
        cell_changes = [
            change
            for operation in session.staged_operations
            for change in operation.cell_changes
        ]
        row_append_count = sum(
            int(operation.metadata.get("row_count", 0))
            for operation in session.staged_operations
            if operation.operation_type == "append_rows"
        )
        preview: dict[str, Any] = {
            "changed_sheets": changed_sheets,
            "changed_ranges": changed_ranges,
            "old_value_count": sum(1 for change in cell_changes if not _is_blank(change.old_value)),
            "new_value_count": sum(1 for change in cell_changes if not _is_blank(change.new_value)),
            "formula_count": sum(1 for change in cell_changes if change.formula),
            "row_append_count": row_append_count,
            "potential_risk_flags": _unique_list(
                [
                    risk
                    for operation in session.staged_operations
                    for risk in operation.risk_flags
                ]
            ),
            "textual_summary": (
                f"{len(session.staged_operations)} staged operation(s), "
                f"{len(changed_sheets)} changed sheet(s), "
                f"{len(cell_changes)} changed cell(s)."
            ),
        }
        if detail_level in {"detailed", "cell_level"}:
            preview["operations"] = [
                operation.to_preview_dict(detail_level) for operation in session.staged_operations
            ]
        if detail_level == "cell_level":
            preview["cell_changes"] = [change.to_dict() for change in cell_changes]

        session.preview_summary = preview
        session.state = "PREVIEWED"
        result = {"preview": preview}
        self._audit(
            operation="preview_edit_session",
            result="success",
            source_id=session.source_id,
            source_mode=session.source_mode,
            target_file=str(session.target_ref.get("path")),
            duration_ms=_duration_ms(started),
            risk_flags=preview["potential_risk_flags"],
        )
        return result

    def validate_edit_session(self, session_id: str) -> dict[str, Any]:
        started = time_module.perf_counter()
        session = self._require_session(session_id)
        source = self._require_local_source(session.source_id)
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        workbook_path = self._resolve_existing_workbook(source, str(session.target_ref["path"]))
        fingerprint = self._file_fingerprint(workbook_path)
        if fingerprint["hash"] != session.base_hash:
            errors.append(
                {
                    "code": "CONFLICT_DETECTED",
                    "message": "Workbook content changed since the session was opened.",
                }
            )
        if int(fingerprint["size"]) != session.base_size:
            warnings.append(
                {
                    "code": "FILE_SIZE_CHANGED",
                    "message": "Workbook size changed since the session was opened.",
                }
            )
        if str(fingerprint["modified_at"]) != session.base_modified_at:
            warnings.append(
                {
                    "code": "MODIFIED_AT_CHANGED",
                    "message": "Workbook modified_at changed since the session was opened.",
                }
            )
        if source.read_only:
            errors.append(
                {
                    "code": "SOURCE_READ_ONLY",
                    "message": f"Source '{source.source_id}' is read-only.",
                }
            )
        if session.read_only and session.staged_operations:
            errors.append(
                {
                    "code": "SESSION_READ_ONLY",
                    "message": "The session is read-only.",
                }
            )
        try:
            workbook = self._load_workbook(session.temp_workbook_path, data_only=False)
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            errors.append(
                {
                    "code": "VALIDATION_FAILED",
                    "message": f"Staged workbook is not openable: {exc}",
                }
            )

        valid = not errors
        result = {"valid": valid, "errors": errors, "warnings": warnings}
        session.validation_summary = result
        if valid:
            session.state = "VALIDATED"
        self._audit(
            operation="validate_edit_session",
            result="success" if valid else "failed",
            source_id=session.source_id,
            source_mode=session.source_mode,
            target_file=str(session.target_ref.get("path")),
            duration_ms=_duration_ms(started),
            risk_flags=[
                error.get("code")
                for error in errors
                if isinstance(error.get("code"), str)
            ],
        )
        return result

    def commit_edit_session(
        self,
        session_id: str,
        *,
        commit_message: str | None = None,
        require_previewed: bool = True,
        require_validated: bool | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        session = self._require_session(session_id, writable=True)
        if not session.staged_operations:
            self._raise_error("VALIDATION_FAILED", "No staged operations were found.")
        if require_previewed and session.preview_summary is None:
            self._raise_error(
                "PREVIEW_REQUIRED",
                "preview_edit_session must be called before commit.",
            )
        source = self._require_local_source(session.source_id)
        workbook_path = self._resolve_existing_workbook(source, str(session.target_ref["path"]))
        fingerprint = self._file_fingerprint(workbook_path)
        if fingerprint["hash"] != session.base_hash:
            self._raise_error(
                "CONFLICT_DETECTED",
                "Workbook changed since the session was opened.",
                detail={
                    "current_hash": fingerprint["hash"],
                    "base_hash": session.base_hash,
                },
            )
        effective_require_validation = (
            self._config.policies.commit_requires_validation
            if require_validated is None
            else require_validated
        )
        if effective_require_validation and (
            session.validation_summary is None or not bool(session.validation_summary.get("valid"))
        ):
            self._raise_error(
                "VALIDATION_FAILED",
                "validate_edit_session must succeed before commit.",
            )

        backup_ref = None
        if not source.read_only:
            backup_ref = self._create_backup(
                source=source,
                workbook_path=workbook_path,
                session=session,
            )

        try:
            shutil.copy2(session.temp_workbook_path, workbook_path)
        except Exception as exc:  # noqa: BLE001
            self._raise_error(
                "COMMIT_FAILED",
                f"Failed to write workbook: {exc}",
                retriable=True,
            )

        final_fingerprint = self._file_fingerprint(workbook_path)
        commit_id = uuid4().hex
        changed_targets = _unique_list(
            [
                f"{operation.sheet}:{operation.affected_range}"
                if operation.sheet and operation.affected_range
                else operation.affected_range or operation.sheet or operation.operation_type
                for operation in session.staged_operations
            ]
        )
        result = {
            "commit_id": commit_id,
            "final_version": final_fingerprint["hash"],
            "backup_ref": backup_ref,
            "changed_targets": changed_targets,
            "commit_message": commit_message,
        }
        session.state = "COMMITTED"
        session.commit_result = result
        session.backup_ref = backup_ref
        self._cleanup_session_files(session)
        self._audit(
            operation="commit_edit_session",
            result="success",
            source_id=session.source_id,
            source_mode=session.source_mode,
            target_file=str(session.target_ref.get("path")),
            duration_ms=_duration_ms(started),
            risk_flags=[
                risk
                for operation in session.staged_operations
                for risk in operation.risk_flags
            ],
            commit_id=commit_id,
        )
        return result

    def cancel_edit_session(self, session_id: str) -> dict[str, Any]:
        started = time_module.perf_counter()
        session = self._require_session(session_id)
        session.state = "CANCELED"
        self._cleanup_session_files(session)
        result = {"canceled": True}
        self._audit(
            operation="cancel_edit_session",
            result="success",
            source_id=session.source_id,
            source_mode=session.source_mode,
            target_file=str(session.target_ref.get("path")),
            duration_ms=_duration_ms(started),
        )
        return result

    def list_backups(
        self,
        source_id: str,
        *,
        target: Mapping[str, Any] | str | None = None,
        limit: int = 50,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        if limit <= 0:
            self._raise_error("VALIDATION_FAILED", "limit must be greater than zero.")
        source = self._require_local_source(source_id)
        target_path = self._normalize_backup_target(target)
        backups: list[dict[str, Any]] = []
        for metadata_path in sorted(
            source.backup_dir.glob("*.json"),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        ):
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            if not isinstance(metadata, dict):
                continue
            if target_path is not None and metadata.get("original_path") != target_path:
                continue
            backups.append(metadata)
            if len(backups) >= limit:
                break
        result = {"backups": backups}
        self._audit(
            operation="list_backups",
            result="success",
            source_id=source.source_id,
            target_file=target_path,
            duration_ms=_duration_ms(started),
        )
        return result

    def restore_backup(
        self,
        backup_ref: Mapping[str, Any] | str,
        *,
        target_override: str | None = None,
    ) -> dict[str, Any]:
        started = time_module.perf_counter()
        metadata = self._resolve_backup_metadata(backup_ref)
        source = self._require_local_source(str(metadata["source_id"]))
        backup_path = self._resolve_path_inside_root(
            str(metadata["backup_path"]),
            source.backup_dir,
        )
        if not backup_path.exists():
            self._raise_error(
                "BACKUP_FAILED",
                f"Backup file '{backup_path.name}' was not found.",
            )
        target_path = (
            self._resolve_path_inside_root(target_override, source.workspace_root)
            if target_override is not None
            else self._resolve_path_inside_root(
                str(metadata["original_path"]),
                source.workspace_root,
            )
        )
        shutil.copy2(backup_path, target_path)
        result = {
            "restore_result": {
                "restored": True,
                "backup_ref": metadata["backup_ref"],
                "target": target_path.relative_to(source.workspace_root).as_posix(),
            }
        }
        self._audit(
            operation="restore_backup",
            result="success",
            source_id=source.source_id,
            target_file=result["restore_result"]["target"],
            duration_ms=_duration_ms(started),
        )
        return result

    # Compatibility methods used by the current planner/test suite.

    def open_file(self, path: str) -> dict[str, Any]:
        workbook_ref = self.resolve_workbook(self._default_local_source().source_id, path=path)[
            "workbook_ref"
        ]
        inspected = self.inspect_workbook(
            workbook_ref,
            include_sheet_stats=False,
            include_tables=False,
        )
        sheet_names = [sheet["name"] for sheet in inspected["sheets"]]
        return {
            "file": workbook_ref["path"],
            "sheet_names": sheet_names,
            "active_sheet": sheet_names[0] if sheet_names else None,
        }

    def read_sheet(self, path: str, sheet: str) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            worksheet = self._get_sheet(workbook, sheet)
            rows: list[dict[str, Any]] = []
            for row in worksheet.iter_rows():
                row_payload = {
                    cell.column_letter: _serialize_value(cell.value)
                    for cell in row
                    if cell.value is not None
                }
                if row_payload:
                    rows.append(row_payload)
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "rows": rows,
            "row_count": len(rows),
        }

    def read_cells(self, path: str, sheet: str, cells: list[str]) -> dict[str, Any]:
        if not cells:
            raise ValueError("cells must contain at least one cell reference.")
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            worksheet = self._get_sheet(workbook, sheet)
            resolved_cells = {
                self._normalize_cell_ref(cell_ref): _serialize_value(
                    worksheet[self._normalize_cell_ref(cell_ref)].value
                )
                for cell_ref in cells
            }
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "cells": resolved_cells,
        }

    def grep_cells(
        self,
        path: str,
        pattern: str,
        *,
        sheet: str | None = None,
        case_sensitive: bool = False,
        regex: bool = False,
        exact: bool = False,
        max_results: int = 100,
    ) -> dict[str, Any]:
        if max_results <= 0:
            raise ValueError("max_results must be greater than zero.")
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            sheet_names = [sheet] if sheet is not None else list(workbook.sheetnames)
            matches: list[dict[str, Any]] = []
            for sheet_name in sheet_names:
                worksheet = self._get_sheet(workbook, sheet_name)
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        if not _matches_text(
                            text=str(cell.value),
                            pattern=pattern,
                            case_sensitive=case_sensitive,
                            regex=regex,
                            exact=exact,
                        ):
                            continue
                        matches.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "value": _serialize_value(cell.value),
                            }
                        )
                        if len(matches) >= max_results:
                            return {
                                "file": workbook_path.relative_to(source.workspace_root).as_posix(),
                                "pattern": pattern,
                                "matches": matches,
                                "truncated": True,
                            }
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "pattern": pattern,
            "matches": matches,
            "truncated": False,
        }

    def calculate_sum(
        self,
        path: str,
        sheet: str,
        column: str,
        start_row: int | None = None,
        end_row: int | None = None,
    ) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=True)
        try:
            worksheet = self._get_sheet(workbook, sheet)
            column_letter = self._normalize_column(column)
            effective_start = start_row or self._infer_start_row(worksheet, column_letter)
            effective_end = end_row or worksheet.max_row
            total = 0.0
            counted_cells = 0
            ignored_cells = 0
            for row_index in range(effective_start, effective_end + 1):
                value = worksheet[f"{column_letter}{row_index}"].value
                numeric = self._coerce_number(value)
                if numeric is None:
                    if value is not None:
                        ignored_cells += 1
                    continue
                total += numeric
                counted_cells += 1
        finally:
            workbook.close()
        normalized_total: int | float = int(total) if total.is_integer() else total
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "column": column_letter,
            "start_row": effective_start,
            "end_row": effective_end,
            "total": normalized_total,
            "counted_cells": counted_cells,
            "ignored_cells": ignored_cells,
        }

    def create_file(
        self,
        path: str,
        sheet: str = "Sheet1",
        overwrite: bool = False,
    ) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_path_inside_root(path, source.workspace_root)
        existed = workbook_path.exists()
        if workbook_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Workbook '{path}' must be an .xlsx file.")
        if existed:
            if workbook_path.is_dir():
                raise IsADirectoryError(f"Workbook path '{path}' is not a file.")
            if not overwrite:
                raise FileExistsError(
                    f"Workbook '{path}' already exists. Set overwrite=True to replace it."
                )
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = self._new_workbook()
        try:
            workbook.active.title = sheet.strip() or "Sheet1"
            workbook.save(workbook_path)
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet_names": [sheet.strip() or "Sheet1"],
            "created": True,
            "overwritten": existed,
        }

    def create_sheet(self, path: str, sheet: str, overwrite: bool = False) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=False)
        try:
            created = False
            if sheet in workbook.sheetnames:
                if overwrite:
                    existing = workbook[sheet]
                    workbook.remove(existing)
                    workbook.create_sheet(title=sheet)
                    created = True
            else:
                workbook.create_sheet(title=sheet)
                created = True
            if created:
                self._create_compat_backup(source, workbook_path)
                workbook.save(workbook_path)
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "created": created,
        }

    def write_cells(self, path: str, sheet: str, cells: dict[str, Any]) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        workbook = self._load_workbook(workbook_path, data_only=False)
        try:
            worksheet = (
                workbook[sheet]
                if sheet in workbook.sheetnames
                else workbook.create_sheet(sheet)
            )
            for cell_ref, value in cells.items():
                if not isinstance(cell_ref, str) or not cell_ref.strip():
                    raise ValueError("cells keys must be non-empty Excel cell references.")
                worksheet[self._normalize_cell_ref(cell_ref)] = value
            self._create_compat_backup(source, workbook_path)
            workbook.save(workbook_path)
        finally:
            workbook.close()
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "written_cells": len(cells),
        }

    def list_images(self, path: str, sheet: str | None = None) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "images": self._load_image_refs(workbook_path, sheet=sheet),
        }

    def extract_image(
        self,
        path: str,
        *,
        sheet: str,
        image_index: int,
        output: str | None = None,
        overwrite: bool = False,
    ) -> dict[str, Any]:
        if image_index <= 0:
            raise ValueError("image_index must be greater than zero.")
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        image_refs = self._load_image_refs(workbook_path, sheet=sheet)
        image_ref = next(
            (
                item
                for item in image_refs
                if item["sheet"] == sheet and item["image_index"] == image_index
            ),
            None,
        )
        if image_ref is None:
            raise KeyError(f"Image index {image_index} was not found on sheet '{sheet}'.")
        if output is None:
            safe_sheet = re.sub(r"[^A-Za-z0-9_.-]+", "_", sheet).strip("_") or "sheet"
            artifact_name = (
                f"{workbook_path.stem}_{safe_sheet}_{image_index}"
                f"{image_ref['extension']}"
            )
            output_path = self._resolve_path_inside_root(
                f".orchestra_artifacts/excel_images/{artifact_name}",
                source.workspace_root,
            )
        else:
            output_path = self._resolve_path_inside_root(output, source.workspace_root)
        if output_path.exists() and not overwrite:
            raise FileExistsError(
                f"Output file '{output_path.relative_to(source.workspace_root).as_posix()}' "
                "already exists. Set overwrite=True to replace it."
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(workbook_path) as archive:
            media_bytes = archive.read(str(image_ref["zip_path"]))
        output_path.write_bytes(media_bytes)
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "sheet": sheet,
            "image_index": image_index,
            "anchor_cell": image_ref["anchor_cell"],
            "output": output_path.relative_to(source.workspace_root).as_posix(),
        }

    def save_file(self, path: str, output: str, overwrite: bool = True) -> dict[str, Any]:
        source = self._default_local_source()
        workbook_path = self._resolve_existing_workbook(source, path)
        output_path = self._resolve_path_inside_root(output, source.workspace_root)
        if output_path.exists() and output_path != workbook_path and not overwrite:
            raise FileExistsError(
                f"Output file '{output}' already exists. Set overwrite=True to replace it."
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if output_path == workbook_path:
            workbook = self._load_workbook(workbook_path, data_only=False)
            try:
                workbook.save(workbook_path)
            finally:
                workbook.close()
        else:
            shutil.copy2(workbook_path, output_path)
        return {
            "file": workbook_path.relative_to(source.workspace_root).as_posix(),
            "output": output_path.relative_to(source.workspace_root).as_posix(),
        }

    def _normalize_sources(
        self,
        sources: Sequence[ExcelSourceProfile],
    ) -> dict[str, ExcelSourceProfile]:
        if not sources:
            fallback = ExcelServerConfig.default(self._workspace_root)
            sources = fallback.sources
        normalized: dict[str, ExcelSourceProfile] = {}
        for source in sources:
            if source.source_id in normalized:
                self._raise_error(
                    "VALIDATION_FAILED",
                    f"Duplicate source_id '{source.source_id}'.",
                )
            if source.source_type == "local_workspace":
                source.workspace_root = self._resolve_inside_workspace(
                    source.workspace_root or self._workspace_root
                )
                source.temp_root = self._resolve_inside_workspace(
                    source.temp_root or (source.workspace_root / ".excel_mcp_tmp")
                )
                source.backup_dir = self._resolve_inside_workspace(
                    source.backup_dir or (source.workspace_root / ".excel_mcp_backups")
                )
                source.temp_root.mkdir(parents=True, exist_ok=True)
                source.backup_dir.mkdir(parents=True, exist_ok=True)
            normalized[source.source_id] = source
        return normalized

    def _resolve_workbook_input(
        self,
        workbook_ref: Mapping[str, Any] | str,
    ) -> tuple[ExcelSourceProfile, Path, dict[str, Any]]:
        if isinstance(workbook_ref, str):
            source = self._default_local_source()
            workbook_path = self._resolve_existing_workbook(source, workbook_ref)
            return source, workbook_path, self._build_workbook_ref(source, workbook_path)
        if not isinstance(workbook_ref, Mapping):
            self._raise_error(
                "WORKBOOK_NOT_FOUND",
                "workbook_ref must be an object or path string.",
            )
        source_id = workbook_ref.get("source_id")
        path = workbook_ref.get("path")
        if not isinstance(source_id, str) or not source_id.strip():
            self._raise_error("SOURCE_NOT_FOUND", "workbook_ref.source_id is required.")
        if not isinstance(path, str) or not path.strip():
            self._raise_error("WORKBOOK_NOT_FOUND", "workbook_ref.path is required.")
        source = self._require_source(source_id)
        if source.source_type != "local_workspace":
            self._raise_remote_not_supported(source)
        workbook_path = self._resolve_existing_workbook(source, path)
        return source, workbook_path, self._build_workbook_ref(source, workbook_path)

    def _require_source(self, source_id: str) -> ExcelSourceProfile:
        source = self._sources.get(source_id)
        if source is None:
            self._raise_error(
                "SOURCE_NOT_FOUND",
                f"Source '{source_id}' was not found.",
            )
        if not source.enabled:
            self._raise_error(
                "SOURCE_DISABLED",
                f"Source '{source_id}' is disabled.",
            )
        return source

    def _require_local_source(self, source_id: str) -> ExcelSourceProfile:
        source = self._require_source(source_id)
        if source.source_type != "local_workspace":
            self._raise_remote_not_supported(source)
        return source

    def _default_local_source(self) -> ExcelSourceProfile:
        for source in self._sources.values():
            if source.source_type == "local_workspace" and source.enabled:
                return source
        self._raise_error("SOURCE_NOT_FOUND", "No enabled local_workspace source is configured.")

    def _resolve_existing_workbook(self, source: ExcelSourceProfile, path: str) -> Path:
        workbook_path = self._resolve_path_inside_root(path, source.workspace_root)
        if not workbook_path.exists():
            self._raise_error(
                "WORKBOOK_NOT_FOUND",
                f"Workbook '{path}' does not exist.",
            )
        if not workbook_path.is_file():
            raise IsADirectoryError(f"Workbook path '{path}' is not a file.")
        self._validate_workbook_extension(path, workbook_path, source)
        self._enforce_file_size_limit(source, workbook_path)
        return workbook_path

    def _resolve_inside_workspace(self, path: Path) -> Path:
        candidate = Path(path)
        resolved = (
            candidate.resolve()
            if candidate.is_absolute()
            else (self._workspace_root / candidate).resolve()
        )
        try:
            resolved.relative_to(self._workspace_root)
        except ValueError as exc:
            raise ValueError(
                f"Path '{resolved}' is outside workspace root '{self._workspace_root}'."
            ) from exc
        return resolved

    def _resolve_path_inside_root(self, relative_path: str | Path, root: Path) -> Path:
        raw = str(relative_path)
        if raw.startswith("\\\\"):
            raise PermissionError("UNC paths are not allowed.")
        candidate = Path(relative_path)
        resolved = candidate.resolve() if candidate.is_absolute() else (root / candidate).resolve()
        try:
            resolved.relative_to(root)
        except ValueError as exc:
            raise PermissionError(
                f"Path '{relative_path}' is outside workspace root '{root}'."
            ) from exc
        return resolved

    def _build_workbook_ref(
        self,
        source: ExcelSourceProfile,
        workbook_path: Path,
        *,
        include_hash: bool = True,
    ) -> dict[str, Any]:
        ref = {
            "source_id": source.source_id,
            "source_type": source.source_type,
            "path": workbook_path.relative_to(source.workspace_root).as_posix(),
            "name": workbook_path.name,
            "size": workbook_path.stat().st_size,
            "modified_at": _file_mtime_iso(workbook_path),
        }
        if include_hash:
            ref["hash"] = _sha256_file(workbook_path)
        return ref

    def _normalize_extension_filter(
        self,
        extension_filter: Sequence[str] | None,
        source: ExcelSourceProfile,
    ) -> set[str]:
        if extension_filter is None:
            return set(source.allowed_extensions)
        normalized: set[str] = set()
        for item in extension_filter:
            lowered = item.lower().strip()
            if not lowered:
                continue
            normalized.add(lowered if lowered.startswith(".") else f".{lowered}")
        return normalized or set(source.allowed_extensions)

    def _validate_workbook_extension(
        self,
        path_label: str,
        workbook_path: Path,
        source: ExcelSourceProfile,
    ) -> None:
        if workbook_path.suffix.lower() not in source.allowed_extensions:
            self._raise_error(
                "WORKBOOK_UNSUPPORTED_FORMAT",
                f"Workbook '{path_label}' must use one of {list(source.allowed_extensions)}.",
            )

    def _enforce_file_size_limit(self, source: ExcelSourceProfile, workbook_path: Path) -> None:
        max_bytes = source.max_file_size_mb * 1024 * 1024
        file_size = workbook_path.stat().st_size
        if file_size > max_bytes:
            self._raise_error(
                "WORKBOOK_TOO_LARGE",
                f"Workbook '{workbook_path.name}' exceeds {source.max_file_size_mb} MB.",
                detail={"size": file_size, "max_bytes": max_bytes},
            )

    def _require_session(self, session_id: str, *, writable: bool = False) -> EditSession:
        session = self._sessions.get(session_id)
        if session is None:
            self._raise_error("SESSION_NOT_FOUND", f"Session '{session_id}' was not found.")
        now = datetime.now(UTC)
        idle_deadline = session.last_accessed_at + timedelta(
            seconds=self._config.limits.idle_timeout_sec
        )
        if now > session.expires_at or now > idle_deadline:
            session.state = "FAILED"
            self._cleanup_session_files(session)
            self._raise_error(
                "SESSION_EXPIRED",
                f"Session '{session_id}' has expired.",
            )
        if session.state in {"COMMITTED", "CANCELED", "FAILED"}:
            self._raise_error(
                "SESSION_NOT_FOUND",
                f"Session '{session_id}' is no longer active.",
            )
        if writable and session.read_only:
            self._raise_error("SESSION_READ_ONLY", "The session is read-only.")
        session.touch()
        return session

    def _record_operation(self, session: EditSession, operation: StagedOperation) -> None:
        session.staged_operations.append(operation)
        session.preview_summary = None
        session.validation_summary = None
        session.state = "STAGING"
        self._audit(
            operation=f"stage_{operation.operation_type}",
            result="success",
            source_id=session.source_id,
            source_mode=session.source_mode,
            target_file=str(session.target_ref.get("path")),
            target_sheet=operation.sheet,
            target_range=operation.affected_range,
            risk_flags=operation.risk_flags,
        )

    def _file_fingerprint(self, path: Path) -> dict[str, Any]:
        stats = path.stat()
        return {
            "hash": _sha256_file(path),
            "size": stats.st_size,
            "modified_at": datetime.fromtimestamp(stats.st_mtime, tz=UTC).isoformat(),
        }

    def _create_backup(
        self,
        *,
        source: ExcelSourceProfile,
        workbook_path: Path,
        session: EditSession,
    ) -> dict[str, Any]:
        timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        backup_filename = (
            f"{timestamp}__{workbook_path.stem}__{session.session_id}"
            f"{workbook_path.suffix}"
        )
        backup_path = source.backup_dir / backup_filename
        try:
            shutil.copy2(workbook_path, backup_path)
        except Exception as exc:  # noqa: BLE001
            self._raise_error(
                "BACKUP_FAILED",
                f"Failed to create backup: {exc}",
                retriable=True,
            )
        metadata = {
            "backup_ref": uuid4().hex,
            "source_id": source.source_id,
            "original_path": workbook_path.relative_to(source.workspace_root).as_posix(),
            "backup_path": backup_path.relative_to(source.backup_dir).as_posix(),
            "sha256": _sha256_file(backup_path),
            "size": backup_path.stat().st_size,
            "session_id": session.session_id,
            "actor": session.actor,
            "created_at": datetime.now(UTC).isoformat(),
        }
        metadata_path = source.backup_dir / f"{metadata['backup_ref']}.json"
        metadata_path.write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        self._audit(
            operation="backup",
            result="success",
            source_id=source.source_id,
            source_mode=session.source_mode,
            target_file=metadata["original_path"],
            duration_ms=None,
        )
        return metadata

    def _create_compat_backup(self, source: ExcelSourceProfile, workbook_path: Path) -> None:
        if not workbook_path.exists():
            return
        session = EditSession(
            session_id=f"compat-{uuid4().hex}",
            source_id=source.source_id,
            source_mode=source.default_mode,
            target_ref={"path": workbook_path.relative_to(source.workspace_root).as_posix()},
            opened_at=datetime.now(UTC),
            expires_at=datetime.now(UTC),
            actor=self._actor(),
            read_only=False,
            base_version="compat",
            base_etag=None,
            base_hash="compat",
            base_size=workbook_path.stat().st_size,
            base_modified_at=_file_mtime_iso(workbook_path),
            temp_workbook_path=workbook_path,
        )
        self._create_backup(source=source, workbook_path=workbook_path, session=session)

    def _resolve_backup_metadata(self, backup_ref: Mapping[str, Any] | str) -> dict[str, Any]:
        reference = (
            backup_ref.get("backup_ref")
            if isinstance(backup_ref, Mapping)
            else backup_ref
        )
        if not isinstance(reference, str) or not reference.strip():
            self._raise_error("BACKUP_FAILED", "backup_ref is required.")
        for source in self._sources.values():
            if source.source_type != "local_workspace":
                continue
            metadata_path = source.backup_dir / f"{reference}.json"
            if metadata_path.exists():
                payload = json.loads(metadata_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    return payload
        self._raise_error("BACKUP_FAILED", f"Backup ref '{reference}' was not found.")

    def _normalize_backup_target(
        self,
        target: Mapping[str, Any] | str | None,
    ) -> str | None:
        if target is None:
            return None
        if isinstance(target, str):
            return target.replace("\\", "/")
        if isinstance(target, Mapping):
            path = target.get("path")
            if isinstance(path, str) and path.strip():
                return path.replace("\\", "/")
        self._raise_error("VALIDATION_FAILED", "target must be a workbook_ref or path string.")

    def _cleanup_session_files(self, session: EditSession) -> None:
        temp_root = session.temp_workbook_path.parent
        shutil.rmtree(temp_root, ignore_errors=True)

    def _raise_remote_not_supported(self, source: ExcelSourceProfile) -> None:
        self._raise_error(
            "REMOTE_API_ERROR",
            (
                f"Remote source '{source.source_id}' ({source.source_type}) is configured, "
                "but this build only executes local_workspace operations."
            ),
            suggested_action="Use a local_workspace source or add a Graph adapter implementation.",
        )

    def _audit(
        self,
        *,
        operation: str,
        result: str,
        source_id: str | None = None,
        source_mode: str | None = None,
        target_file: str | None = None,
        target_sheet: str | None = None,
        target_range: str | None = None,
        duration_ms: int | None = None,
        risk_flags: Sequence[str] | None = None,
        commit_id: str | None = None,
    ) -> None:
        payload = {
            "timestamp": datetime.now(UTC).isoformat(),
            "request_id": uuid4().hex,
            "actor": self._actor(),
            "source_id": source_id,
            "source_mode": source_mode,
            "operation": operation,
            "target_file": target_file,
            "target_sheet": target_sheet,
            "target_range": target_range,
            "result": result,
            "duration_ms": duration_ms,
            "risk_flags": list(risk_flags or []),
            "commit_id": commit_id,
            "remote_item_id": None,
            "remote_etag_before": None,
            "remote_etag_after": None,
        }
        self._audit_file.parent.mkdir(parents=True, exist_ok=True)
        with self._audit_file.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(payload, ensure_ascii=False))
            handle.write("\n")

    @staticmethod
    def _actor() -> str:
        for env_name in ("EXCEL_MCP_ACTOR", "USERNAME", "USER"):
            value = os.getenv(env_name)
            if value and value.strip():
                return value.strip()
        return "unknown"

    @staticmethod
    def _load_workbook(path: Path, data_only: bool) -> Any:
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ImportError(
                "Missing dependency 'openpyxl'. Install optional extras with "
                "`pip install \"orchestra-agent[mcp-server]\"`."
            ) from exc
        return load_workbook(
            filename=path,
            data_only=data_only,
            keep_vba=path.suffix.lower() == ".xlsm",
        )

    @staticmethod
    def _new_workbook() -> Any:
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ImportError(
                "Missing dependency 'openpyxl'. Install optional extras with "
                "`pip install \"orchestra-agent[mcp-server]\"`."
            ) from exc
        return Workbook()

    @staticmethod
    def _get_sheet(workbook: Any, sheet: str) -> Any:
        if sheet not in workbook.sheetnames:
            raise KeyError(f"Worksheet '{sheet}' does not exist.")
        return workbook[sheet]

    @staticmethod
    def _normalize_column(column: str) -> str:
        normalized = column.strip().upper()
        if not re.fullmatch(r"[A-Z]{1,3}", normalized):
            raise ValueError(f"Invalid Excel column reference: '{column}'.")
        return normalized

    @staticmethod
    def _normalize_cell_ref(cell_ref: str) -> str:
        normalized = cell_ref.strip().upper()
        if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]*", normalized):
            raise ValueError(f"Invalid Excel cell reference: '{cell_ref}'.")
        return normalized

    def _cell_indices(self, cell_ref: str) -> tuple[int, int]:
        from openpyxl.utils.cell import (  # type: ignore[import-untyped]
            column_index_from_string,
            coordinate_from_string,
        )

        normalized = self._normalize_cell_ref(cell_ref)
        column_letters, row_index = coordinate_from_string(normalized)
        return column_index_from_string(column_letters), row_index

    @staticmethod
    def _range_boundaries(range_ref: str) -> tuple[int, int, int, int]:
        from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]

        try:
            return range_boundaries(range_ref.upper())
        except ValueError as exc:
            raise ExcelToolError("RANGE_INVALID", f"Invalid range '{range_ref}'.") from exc

    def _find_table(
        self,
        workbook: Any,
        table_name: str,
        *,
        sheet: str | None = None,
    ) -> tuple[Any, Any]:
        worksheets = (
            [self._get_sheet(workbook, sheet)]
            if sheet is not None
            else workbook.worksheets
        )
        for worksheet in worksheets:
            table = worksheet.tables.get(table_name)
            if table is not None:
                return worksheet, table
        self._raise_error(
            "TABLE_NOT_FOUND",
            f"Table '{table_name}' does not exist.",
        )

    def _render_cell_value(self, value: Any, mode: ValueRenderMode) -> Any:
        if mode == "formatted":
            if value is None:
                return None
            if isinstance(value, bool):
                return "TRUE" if value else "FALSE"
            if isinstance(value, datetime | date | time_value):
                return value.isoformat()
            return str(value)
        return _serialize_value(value)

    def _ensure_not_merged_overlap(self, worksheet: Any, cell_ref: str) -> None:
        for merged_range in worksheet.merged_cells.ranges:
            if cell_ref in merged_range and cell_ref != merged_range.start_cell.coordinate:
                self._raise_error(
                    "VALIDATION_FAILED",
                    f"Cell {cell_ref} overlaps a merged cell range.",
                    detail={"cell": cell_ref, "merged_range": str(merged_range)},
                )

    @staticmethod
    def _infer_start_row(worksheet: Any, column_letter: str) -> int:
        first_value = worksheet[f"{column_letter}1"].value
        if first_value is None:
            return 1
        if ExcelWorkspaceService._coerce_number(first_value) is None:
            return 2
        return 1

    @staticmethod
    def _coerce_number(value: Any) -> float | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, int | float):
            return float(value)
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    def _load_image_refs(
        self,
        workbook_path: Path,
        *,
        sheet: str | None = None,
    ) -> list[dict[str, Any]]:
        with ZipFile(workbook_path) as archive:
            sheet_targets = self._sheet_targets(archive)
            if sheet is not None and sheet not in sheet_targets:
                raise KeyError(f"Worksheet '{sheet}' does not exist.")
            image_refs: list[dict[str, Any]] = []
            for sheet_name, sheet_xml_path in sheet_targets.items():
                if sheet is not None and sheet_name != sheet:
                    continue
                image_refs.extend(
                    self._sheet_image_refs(
                        archive=archive,
                        sheet_name=sheet_name,
                        sheet_xml_path=sheet_xml_path,
                    )
                )
        return image_refs

    def _sheet_targets(self, archive: ZipFile) -> dict[str, str]:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = self._relationships(archive, "xl/_rels/workbook.xml.rels")
        sheet_targets: dict[str, str] = {}
        sheet_path = f"{{{self._main_ns}}}sheets/{{{self._main_ns}}}sheet"
        for sheet_elem in workbook_root.findall(sheet_path):
            sheet_name = sheet_elem.attrib.get("name")
            rel_id = sheet_elem.attrib.get(f"{{{self._rel_ns}}}id")
            if not isinstance(sheet_name, str) or not isinstance(rel_id, str):
                continue
            target = workbook_rels.get(rel_id)
            if target is None:
                continue
            sheet_targets[sheet_name] = self._resolve_zip_target("xl/workbook.xml", target)
        return sheet_targets

    def _sheet_image_refs(
        self,
        *,
        archive: ZipFile,
        sheet_name: str,
        sheet_xml_path: str,
    ) -> list[dict[str, Any]]:
        if sheet_xml_path not in archive.namelist():
            return []
        sheet_root = ET.fromstring(archive.read(sheet_xml_path))
        sheet_rels = self._relationships(archive, self._rels_path(sheet_xml_path))
        drawing_rel_ids = [
            elem.attrib.get(f"{{{self._rel_ns}}}id")
            for elem in sheet_root.findall(f"{{{self._main_ns}}}drawing")
        ]
        image_refs: list[dict[str, Any]] = []
        image_index = 0
        for drawing_rel_id in drawing_rel_ids:
            if not isinstance(drawing_rel_id, str):
                continue
            drawing_target = sheet_rels.get(drawing_rel_id)
            if drawing_target is None:
                continue
            drawing_xml_path = self._resolve_zip_target(sheet_xml_path, drawing_target)
            refs_from_drawing = self._drawing_image_refs(
                archive=archive,
                sheet_name=sheet_name,
                drawing_xml_path=drawing_xml_path,
                image_index_offset=image_index,
            )
            image_refs.extend(refs_from_drawing)
            image_index += len(refs_from_drawing)
        return image_refs

    def _drawing_image_refs(
        self,
        *,
        archive: ZipFile,
        sheet_name: str,
        drawing_xml_path: str,
        image_index_offset: int,
    ) -> list[dict[str, Any]]:
        if drawing_xml_path not in archive.namelist():
            return []
        drawing_root = ET.fromstring(archive.read(drawing_xml_path))
        drawing_rels = self._relationships(archive, self._rels_path(drawing_xml_path))
        image_refs: list[dict[str, Any]] = []
        for anchor in drawing_root:
            embed_id = self._embedded_image_rel_id(anchor)
            if embed_id is None:
                continue
            media_target = drawing_rels.get(embed_id)
            if media_target is None:
                continue
            media_path = self._resolve_zip_target(drawing_xml_path, media_target)
            image_refs.append(
                {
                    "sheet": sheet_name,
                    "image_index": image_index_offset + len(image_refs) + 1,
                    "anchor_cell": self._anchor_cell(anchor),
                    "extension": Path(media_path).suffix.lower(),
                    "zip_path": media_path,
                }
            )
        return image_refs

    def _relationships(self, archive: ZipFile, rels_path: str) -> dict[str, str]:
        if rels_path not in archive.namelist():
            return {}
        rels_root = ET.fromstring(archive.read(rels_path))
        relationships: dict[str, str] = {}
        for rel_elem in rels_root.findall(f"{{{self._pkg_rel_ns}}}Relationship"):
            rel_id = rel_elem.attrib.get("Id")
            target = rel_elem.attrib.get("Target")
            if isinstance(rel_id, str) and isinstance(target, str):
                relationships[rel_id] = target
        return relationships

    @staticmethod
    def _rels_path(xml_path: str) -> str:
        xml_file = PurePosixPath(xml_path)
        return str(xml_file.parent / "_rels" / f"{xml_file.name}.rels")

    @staticmethod
    def _resolve_zip_target(base_xml_path: str, target: str) -> str:
        target_path = PurePosixPath(target)
        if target.startswith("/"):
            return str(PurePosixPath(target.lstrip("/")))
        normalized_parts: list[str] = []
        for part in (PurePosixPath(base_xml_path).parent / target_path).parts:
            if part in ("", "."):
                continue
            if part == "..":
                if normalized_parts:
                    normalized_parts.pop()
                continue
            normalized_parts.append(part)
        return str(PurePosixPath(*normalized_parts))

    def _embedded_image_rel_id(self, anchor: ET.Element) -> str | None:
        blip = anchor.find(f".//{{{self._a_ns}}}blip")
        if blip is None:
            return None
        embed_id = blip.attrib.get(f"{{{self._rel_ns}}}embed")
        return embed_id if isinstance(embed_id, str) else None

    def _anchor_cell(self, anchor: ET.Element) -> str | None:
        from_elem = anchor.find(f"{{{self._xdr_ns}}}from")
        if from_elem is None:
            return None
        row_elem = from_elem.find(f"{{{self._xdr_ns}}}row")
        col_elem = from_elem.find(f"{{{self._xdr_ns}}}col")
        if row_elem is None or col_elem is None:
            return None
        row_index = int(row_elem.text or "0") + 1
        col_index = int(col_elem.text or "0")
        return f"{self._column_letters(col_index)}{row_index}"

    @staticmethod
    def _column_letters(index: int) -> str:
        if index < 0:
            raise ValueError("Excel column index must be zero or greater.")
        value = index + 1
        letters = ""
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            letters = chr(ord("A") + remainder) + letters
        return letters

    @staticmethod
    def _raise_error(code: str, message: str, **kwargs: Any) -> None:
        raise ExcelToolError(code, message, **kwargs)


def _worksheet_is_empty(worksheet: Any) -> bool:
    return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None


def _normalize_matrix(values: Sequence[Sequence[Any]], *, field_name: str) -> list[list[Any]]:
    if not isinstance(values, Sequence) or isinstance(values, str | bytes | bytearray):
        raise ExcelToolError("VALIDATION_FAILED", f"{field_name} must be a 2D array.")
    rows: list[list[Any]] = []
    for row in values:
        if not isinstance(row, Sequence) or isinstance(row, str | bytes | bytearray):
            raise ExcelToolError("VALIDATION_FAILED", f"{field_name} must be a 2D array.")
        rows.append(list(row))
    if not rows or not rows[0]:
        raise ExcelToolError("VALIDATION_FAILED", f"{field_name} must not be empty.")
    expected_width = len(rows[0])
    if expected_width == 0:
        raise ExcelToolError("VALIDATION_FAILED", f"{field_name} rows must not be empty.")
    if any(len(row) != expected_width for row in rows):
        raise ExcelToolError("VALIDATION_FAILED", f"{field_name} must be rectangular.")
    return rows


def _matrix_shape(values: Sequence[Sequence[Any]]) -> tuple[int, int]:
    return len(values), len(values[0]) if values else 0


def _cell_range(*, start_col: int, start_row: int, end_col: int, end_row: int) -> str:
    from openpyxl.utils.cell import get_column_letter  # type: ignore[import-untyped]

    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _matrix_range(start_cell: str, *, row_count: int, col_count: int) -> str:
    from openpyxl.utils.cell import (  # type: ignore[import-untyped]
        column_index_from_string,
        coordinate_from_string,
        get_column_letter,
    )

    column_letters, start_row = coordinate_from_string(start_cell.upper())
    start_col = column_index_from_string(column_letters)
    end_col = start_col + col_count - 1
    end_row = start_row + row_count - 1
    return f"{column_letters}{start_row}:{get_column_letter(end_col)}{end_row}"


def _serialize_value(value: Any) -> Any:
    if value is None or isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date | time_value):
        return value.isoformat()
    return str(value)


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_blank(value: Any) -> bool:
    return value is None or value == ""


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _file_mtime_iso(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=UTC).isoformat()


def _duration_ms(started: float) -> int:
    return int((time_module.perf_counter() - started) * 1000)


def _unique_list(values: Sequence[Any]) -> list[Any]:
    seen: set[Any] = set()
    result: list[Any] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _matches_text(
    text: str,
    pattern: str,
    *,
    case_sensitive: bool,
    regex: bool,
    exact: bool,
) -> bool:
    if regex:
        flags = 0 if case_sensitive else re.IGNORECASE
        return re.search(pattern, text, flags=flags) is not None
    left = text if case_sensitive else text.lower()
    right = pattern if case_sensitive else pattern.lower()
    if exact:
        return left == right
    return right in left
